import openpyxl
import pyperclip
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Function to find the next empty cell in a column
def find_next_empty_cell(sheet, col):
    row = 1
    while sheet.cell(row=row, column=col).value is not None:
        row += 1
    return row

# Function to update Excel sheet with clipboard content
def update_excel():
    # Load or create workbook
    try:
        wb = openpyxl.load_workbook('data1.xlsx')
    except FileNotFoundError:
        wb = Workbook()

    # Select active sheet
    sheet = wb.active
    
    # Get current clipboard content
    clipboard_text = pyperclip.paste().strip()
    
    if clipboard_text:
        # Find next available cell in column A
        next_row = find_next_empty_cell(sheet, 1)
        
        # Update cell with clipboard content
        sheet.cell(row=next_row, column=1).value = clipboard_text

        # Save the workbook
        wb.save('data1.xlsx')
        print(f'Added "{clipboard_text}" to cell A{next_row}')

# Monitor clipboard changes
while True:
    current_clipboard = pyperclip.paste()
    
    # Wait until clipboard content changes
    while pyperclip.paste() == current_clipboard:
        pass
    
    # Update Excel with new clipboard content
    update_excel()